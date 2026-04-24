#!/usr/bin/env python3
"""
run_validate.py
Multi-model dashboard runner with MTDF formula engine.

Author: Ingo Mesche
Affiliation: Independent Researcher, Malta
Date: December 2025
Source of truth: DB_Workbook_STRICT_V18.xlsx, Validation_Dashboard_V74.html

Usage:
  python run_validate.py --workbook Data/DB_Workbook_STRICT_V18.xlsx --out Validation_Dashboard_V74.html --diag Diagnostics.csv
"""
import json
__RV_VERSION__ = "RV74-2025-12"

import argparse
import datetime as _dt
from pathlib import Path
import pandas as pd
import math
import sys
import re
from types import SimpleNamespace

# === PATCH: SPARC helpers (keep above any prints) ============================
import os, json
import numpy as np

# meters per kiloparsec (define if not already defined)
KPC_M = globals().get("KPC_M", 3.085677581e19)

# Resolve SPARC data path relative to this script (not cwd)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_SPARC_PATH = os.path.join(_SCRIPT_DIR, "..", "data", "sparc_clean.json")

# Vector pillar support
try:
    from vector_pillars import (
        load_pantheonplus, load_desi_bao, load_cc_hz, load_dr16_fsigma8,
        load_cmb_distance_prior,
        mtdf_mu_vector, mtdf_bao_vector, mtdf_Hz_vector, mtdf_fsigma8_vector,
        mtdf_cmb_distance_vector,
        lcdm_fsigma8_vector, fit_sigma8_lcdm,
        chi2_vector_pillar, chi2_sne_marginalized, evaluate_vector_pillar,
        get_vector_pillar_configs
    )
    VECTOR_PILLARS_AVAILABLE = True
    print("[VECTOR] Vector pillar module loaded successfully")
except ImportError as e:
    VECTOR_PILLARS_AVAILABLE = False
    print(f"[VECTOR] Vector pillar module not available: {e}")

# Vector pillar EFE support (MTDF with Early Field Energy)
try:
    from vector_pillars_EFE import (
        mtdf_cmb_distance_vector_efe,
    )
    VECTOR_PILLARS_EFE_AVAILABLE = True
    print("[VECTOR] Vector pillar EFE module loaded successfully")
except ImportError as e:
    VECTOR_PILLARS_EFE_AVAILABLE = False
    print(f"[VECTOR] Vector pillar EFE module not available: {e}")

def _load_sparc_velocity_fields(path=_SPARC_PATH):
    """
    Returns r (meters), v_obs_kms, v_bar_kms, v_err_kms from SPARC JSON.
    If v_err is missing, returns zeros.
    """
    if not os.path.exists(path):
        return None, None, None, None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None, None, None, None

    r_kpc_all, v_obs_all, v_bar_all, v_err_all = [], [], [], []
    galaxies = data.get("galaxies") or {}
    for gal in galaxies.values():
        pts = gal.get("points") or []
        r_kpc, vobs, vbar, verr = [], [], [], []
        for p in pts:
            r = p.get("r_kpc") or p.get("r")
            vobs_i = p.get("v_obs", p.get("v_obs_kms"))
            vbar_i = p.get("v_bar", p.get("v_bar_kms"))
            verr_i = p.get("v_err", p.get("v_err_kms"))
            if r is None or vobs_i is None or vbar_i is None:
                continue
            try:
                r_kpc.append(float(r))
                vobs.append(float(vobs_i))
                vbar.append(float(vbar_i))
                verr.append(float(verr_i) if verr_i is not None else 0.0)
            except Exception:
                continue
        n = min(len(r_kpc), len(vobs), len(vbar), len(verr))
        if n < 3:
            continue
        r_kpc_all.extend(r_kpc[:n])
        v_obs_all.extend(vobs[:n])
        v_bar_all.extend(vbar[:n])
        v_err_all.extend(verr[:n])

    if not r_kpc_all:
        return None, None, None, None

    r_m = np.asarray(r_kpc_all) * KPC_M
    v_obs_kms = np.asarray(v_obs_all)
    v_bar_kms = np.asarray(v_bar_all)
    v_err_kms = np.asarray(v_err_all)
    return r_m, v_obs_kms, v_bar_kms, v_err_kms


def _load_sparc_radii(path=_SPARC_PATH):
    """
    Returns concatenated radii array in meters from SPARC JSON.
    """
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None
    r_kpc_all = []
    galaxies = data.get("galaxies") or {}
    for gal in galaxies.values():
        pts = gal.get("points") or []
        for p in pts:
            r = p.get("r_kpc") or p.get("r")
            try:
                if r is not None:
                    r_kpc_all.append(float(r))
            except Exception:
                continue
    if not r_kpc_all:
        return None
    return np.asarray(r_kpc_all) * KPC_M

# Show we DEFINITELY have the helpers in scope now
print("[CHECK] helpers:", "_load_sparc_velocity_fields" in globals(), "_load_sparc_radii" in globals())
# ============================================================================ 


def _load_sparc_profiles(path=_SPARC_PATH):
    """
    Returns flat arrays (g_obs_profile, g_bar_profile) in m/s^2 from a SPARC JSON.
    Updated to handle the actual SPARC data format with 'r', 'v_obs', 'v_bar' fields.
    Gracefully returns (None, None) if file or fields are missing.
    """
    if not os.path.exists(path):
        return None, None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None, None

    g_obs_list, g_bar_list = [], []
    galaxies = data.get("galaxies") or {}
    
    for _, gal in galaxies.items():
        pts = gal.get("points") or []
        r_kpc, v_obs_kms, v_bar_kms = [], [], []
        
        for p in pts:
            # Handle both possible field names for radius
            r = p.get("r_kpc") or p.get("r")  # Try r_kpc first, then r
            # Handle both possible field names for velocities
            vobs = p.get("v_obs", p.get("v_obs_kms"))
            vbar = p.get("v_bar", p.get("v_bar_kms"))
            
            if r is None or vobs is None or vbar is None:
                continue
                
            try:
                r_kpc.append(float(r))
                v_obs_kms.append(float(vobs))
                v_bar_kms.append(float(vbar))
            except (ValueError, TypeError):
                continue
                
        n = min(len(r_kpc), len(v_obs_kms), len(v_bar_kms))
        if n < 3:
            continue
            
        r_m = np.asarray(r_kpc[:n]) * KPC_M
        g_obs = (np.asarray(v_obs_kms[:n]) * 1e3) ** 2 / np.clip(r_m, 1e-6, None)
        g_bar = (np.asarray(v_bar_kms[:n]) * 1e3) ** 2 / np.clip(r_m, 1e-6, None)
        g_obs_list.append(g_obs)
        g_bar_list.append(g_bar)

    if not g_obs_list:
        return None, None
    return np.concatenate(g_obs_list), np.concatenate(g_bar_list)

def _rar_intrinsic_scatter(g_obs=None, g_bar=None, errors="SPARC", get_token=None, **kwargs):
    print(f"[DEBUG] _rar_intrinsic_scatter called with get_token={get_token is not None}")
    print(f"[DEBUG] g_obs size: {g_obs.size if g_obs is not None else None}")
    
    """
    Calculate intrinsic scatter of the RAR after deconvolving measurement errors.
    Based on Desmond 2023 methodology for SPARC data.
    
    Args:
        g_obs: Observed acceleration array (m/s^2)
        g_bar: Baryonic acceleration array (m/s^2) 
        errors: Error treatment method ('SPARC' uses SPARC-specific parameters)
        get_token: Function to retrieve parameter values
        
    Returns:
        float: Intrinsic scatter in dex after error deconvolution
    """
    if g_obs is None or g_bar is None:
        return float("nan")
    
    go = np.asarray(g_obs, dtype=float)
    gb = np.asarray(g_bar, dtype=float)
    
    if go.shape != gb.shape or go.size < 3:
        return float("nan")
    
    # Calculate log residuals
    log_g_obs = np.log10(np.clip(go, 1e-30, None))
    log_g_bar = np.log10(np.clip(gb, 1e-30, None))
    residuals = log_g_obs - log_g_bar
    
    # Calculate observed scatter
    sigma_obs = float(np.nanstd(residuals))
    print(f"[DEBUG] sigma_obs (raw observed scatter): {sigma_obs}")
    
    # Error deconvolution based on parameters
    if errors == "SPARC" and get_token is not None:
        sigma_meas = get_token("sparc_meas_uncertainty")
        correction_factor = get_token("rar_deconv_factor")
        print(f"[DEBUG] Retrieved: sigma_meas={sigma_meas}, correction_factor={correction_factor}")
        
        if sigma_meas is None:
            sigma_meas = 0.12  # fallback
        if correction_factor is None:
            correction_factor = 0.85  # fallback
            
        # Simple quadrature deconvolution
        # σ_int^2 = σ_obs^2 - σ_meas^2
        sigma_intrinsic_sq = max(0.0, sigma_obs**2 - float(sigma_meas)**2)
        sigma_intrinsic = np.sqrt(sigma_intrinsic_sq)
        
        print(f"[DEBUG] sigma_obs^2: {sigma_obs**2}")
        print(f"[DEBUG] sigma_meas^2: {float(sigma_meas)**2}")
        print(f"[DEBUG] sigma_intrinsic_sq: {sigma_intrinsic_sq}")
        print(f"[DEBUG] sigma_intrinsic: {sigma_intrinsic}")
        
        result = float(sigma_intrinsic * float(correction_factor))
        print(f"[DEBUG] final result after correction: {result}")
        return result
            
    else:
        # Fallback: return observed scatter without deconvolution
        return sigma_obs

def _load_vector_pillar_config(workbook_path):
    """
    Load vector pillar configuration from Pillar_Vector_Config sheet.

    Returns list of config dicts for active pillars.
    """
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if 'Pillar_Vector_Config' not in wb.sheetnames:
        print("[VECTOR] Pillar_Vector_Config sheet not found - using defaults")
        wb.close()
        return None

    ws = wb['Pillar_Vector_Config']

    # Find header row (row 2)
    headers = [cell.value for cell in ws[2]]

    configs = []
    for row_idx in range(3, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, len(headers) + 1)]
        row_dict = dict(zip(headers, row_vals))

        # Skip inactive pillars
        if row_dict.get('active') != 1:
            continue

        configs.append({
            'pillar_id': row_dict.get('pillar_id'),
            'pillar_name': row_dict.get('pillar_name'),
            'category': (row_dict.get('category') or '').upper(),
            'dataset': row_dict.get('dataset'),
            'pillar_mode': row_dict.get('pillar_mode', 'VECTOR'),
            'data_file': row_dict.get('data_file'),
            'cov_file': row_dict.get('cov_file'),
            'model_function': row_dict.get('model_function'),
            'nuisance_count': int(row_dict.get('nuisance_count', 0) or 0),
            'notes': row_dict.get('notes'),
        })

    wb.close()
    return configs


def _evaluate_vector_pillars(get_token, data_dir="data", workbook_path=None, use_efe_cmb=False):
    """
    Evaluate all vector pillars and return combined statistics.

    Reads pillar configuration from workbook Pillar_Vector_Config sheet if available.

    Args:
        use_efe_cmb: If True, use mtdf_cmb_distance_vector_efe (MTDF Early Field Energy)
                     for CMB pillar instead of baseline mtdf_cmb_distance_vector.

    Returns dict with:
        - vector_results: list of individual pillar results
        - total_chi2: sum of all vector pillar chi²
        - total_dof: sum of all vector pillar DOF
        - combined_chi2_red: total_chi2 / total_dof
    """
    if not VECTOR_PILLARS_AVAILABLE:
        return {
            "vector_results": [],
            "total_chi2": 0.0,
            "total_dof": 0,
            "combined_chi2_red": float("nan"),
        }

    # Build params dict from token store
    # -------------------------------------------------------------------------
    # PARAMETER HIERARCHY:
    #   Primary (external microphysics):
    #     omegab_h2 = Omega_b * h^2  (BBN + CMB constraint, fixed)
    #     omegam_h2 = Omega_m * h^2  (physical matter density)
    #   MTDF choice:
    #     H0 (determines h = H0/100)
    #   Derived:
    #     Omega_b = omegab_h2 / h^2
    #     Omega_m = omegam_h2 / h^2
    # -------------------------------------------------------------------------

    # MTDF Hubble parameter (this is the MTDF calibration choice)
    H0 = get_token("h0") or get_token("hubble_constant") or 70.0
    h = H0 / 100.0

    # External microphysics constraints (these are fixed by BBN/CMB)
    omegab_h2 = get_token("omegab_h2") or get_token("omega_b_h2") or 0.02236  # Planck 2018
    omegam_h2 = get_token("omegam_h2") or get_token("omega_m_h2") or 0.1430   # ~Planck physical matter

    # Derive Omega_b and Omega_m from physical densities and H0
    Omega_b_derived = omegab_h2 / h**2
    Omega_m_derived = omegam_h2 / h**2

    # Allow workbook override for Omega_m if explicitly set, but Omega_b is ALWAYS derived
    Omega_m_override = get_token("omega_m") or get_token("omega_matter")
    if Omega_m_override is not None:
        Omega_m = Omega_m_override
        print(f"[VECTOR] Note: Using workbook Ω_m={Omega_m:.4f} (derived would be {Omega_m_derived:.4f})")
    else:
        Omega_m = Omega_m_derived

    params = {
        "H0": H0,
        "h": h,
        # Primary physical densities (external constraints)
        "omegab_h2": omegab_h2,
        "omegam_h2": omegam_h2,
        # Derived density parameters
        "Omega_b": Omega_b_derived,  # ALWAYS derived from omegab_h2
        "Omega_m": Omega_m,
        # MTDF parameters
        "kappa": get_token("kappa") or get_token("kappa_spec") or 1.02e-3,  # Anchor: approx f_kick/3 = (1-beta_eos)^2/(72*(1+alpha))
        "alpha": get_token("alpha") or 1.30,
        "beta_eos": get_token("beta_eos") or get_token("betaeos") or 0.573,
        "z_t": get_token("z_t") or get_token("zt") or 0.74,
        "r_d": get_token("r_d") or get_token("sound_horizon") or 147.09,
        "sigma8": get_token("sigma8") or get_token("sigma_8") or 0.811,
        "gamma_growth": get_token("gamma_growth") or 0.55,
    }
    print(f"[VECTOR] Primary: ωb h²={params['omegab_h2']:.5f}, ωm h²={params['omegam_h2']:.4f}, H0={params['H0']}")
    print(f"[VECTOR] Derived: Ω_b={params['Omega_b']:.5f}, Ω_m={params['Omega_m']:.4f}")
    print(f"[VECTOR]   MTDF: α={params['alpha']}, β_eos={params['beta_eos']}, z_t={params['z_t']}")

    # Load pillar configs from workbook if available
    pillar_configs = None
    if workbook_path:
        pillar_configs = _load_vector_pillar_config(workbook_path)

    # Fallback to hardcoded defaults if workbook config not available
    if not pillar_configs:
        pillar_configs = [
            {"pillar_id": "P_SNE_PANTHEON", "pillar_name": "Pantheon+ SNe", "category": "SNE", "nuisance_count": 1},
            {"pillar_id": "P_BAO_DESI", "pillar_name": "DESI BAO", "category": "BAO", "nuisance_count": 0},
            {"pillar_id": "P_HZ_CC", "pillar_name": "CC H(z)", "category": "HZ", "nuisance_count": 0},
            {"pillar_id": "P_GROWTH_FSIG8", "pillar_name": "DR16 fσ₈", "category": "GROWTH", "nuisance_count": 1},
        ]
        print("[VECTOR] Using default pillar configs (no workbook sheet)")
    else:
        print(f"[VECTOR] Loaded {len(pillar_configs)} vector pillars from workbook")

    results = []
    total_chi2 = 0.0
    total_dof = 0

    for config in pillar_configs:
        try:
            category = config.get('category', '').upper()
            pillar_id = config.get('pillar_id', 'Unknown')
            pillar_name = config.get('pillar_name', pillar_id)
            nuisance_count = config.get('nuisance_count', 0)

            result = {
                "pillar_id": pillar_id,
                "name": pillar_name,
                "category": category,
                "mode": "VECTOR",
            }

            if category == "SNE":
                z, mu_obs, cov = load_pantheonplus(data_dir)
                chi2, dof = chi2_sne_marginalized(z, mu_obs, cov, params)
                result["n_data"] = len(z)
                # SNe marginalization already accounts for M, so dof is correct

            elif category == "BAO":
                z_eff, obs_vec, obs_types, cov = load_desi_bao(data_dir)
                model_pred = mtdf_bao_vector(z_eff, obs_types, params)
                chi2, dof = chi2_vector_pillar(obs_vec, model_pred, cov, nuisance_count)
                result["n_data"] = len(obs_vec)

            elif category == "HZ":
                z, H_obs, cov = load_cc_hz(data_dir)
                model_pred = mtdf_Hz_vector(z, params)
                chi2, dof = chi2_vector_pillar(H_obs, model_pred, cov, nuisance_count)
                result["n_data"] = len(z)

            elif category == "GROWTH":
                z, fsig8_obs, cov = load_dr16_fsigma8(data_dir)
                # Fit σ₈,₀ analytically to minimize χ²
                model_pred, fsig8_diagnostics = mtdf_fsigma8_vector(
                    z, params, return_diagnostics=True,
                    fit_sigma8=True, fsig8_obs=fsig8_obs, cov_matrix=cov
                )
                # Use fitted χ² and DOF
                chi2 = fsig8_diagnostics.get('chi2_fitted', 0.0)
                dof = len(z) - nuisance_count
                result["n_data"] = len(z)
                result["fsig8_obs"] = fsig8_obs
                result["fsig8_diagnostics"] = fsig8_diagnostics
                result["sigma8_bf"] = fsig8_diagnostics.get('sigma8_bf')
                result["sigma8_err"] = fsig8_diagnostics.get('sigma8_err')

            elif category == "CMB":
                obs_means, cov = load_cmb_distance_prior(data_dir)
                # Use EFE CMB function if requested and available
                if use_efe_cmb and VECTOR_PILLARS_EFE_AVAILABLE:
                    model_pred, cmb_diag = mtdf_cmb_distance_vector_efe(params, return_diagnostics=True)
                    result["efe_mode"] = True
                    print(f"[VECTOR] CMB using EFE (Early Field Energy) correction")
                else:
                    model_pred, cmb_diag = mtdf_cmb_distance_vector(params, return_diagnostics=True)
                    result["efe_mode"] = False
                chi2, dof = chi2_vector_pillar(obs_means, model_pred, cov, nuisance_count)
                result["n_data"] = len(obs_means)
                result["cmb_diagnostics"] = cmb_diag
                result["cmb_obs"] = obs_means
                result["cmb_pred"] = model_pred

            else:
                print(f"[VECTOR] Unknown category '{category}' for {pillar_name}, skipping")
                continue

            result["chi2"] = chi2
            result["dof"] = dof
            result["chi2_red"] = chi2 / dof if dof > 0 else float("nan")
            result["z_equiv"] = (chi2 - dof) / np.sqrt(2 * dof) if dof > 0 else float("nan")
            result["experimental"] = False  # All workbook pillars are production

            results.append(result)

            # Add to combined totals
            total_chi2 += chi2
            total_dof += dof

            print(f"[VECTOR] {pillar_name}: n={result['n_data']}, χ²={chi2:.1f}, DOF={dof}, χ²/ν={result['chi2_red']:.4f}")

        except Exception as e:
            print(f"[VECTOR] Error evaluating {config.get('pillar_name', 'Unknown')}: {e}")
            import traceback
            traceback.print_exc()

    return {
        "vector_results": results,
        "total_chi2": total_chi2,
        "total_dof": total_dof,
        "combined_chi2_red": total_chi2 / total_dof if total_dof > 0 else float("nan"),
    }


def _norm_doi(s):
    if s is None:
        return ""
    s = str(s)
    m = re.search(r"(10\.\d{4,9}/\S+)", s)
    return f"https://doi.org/{m.group(1)}" if m else ""

def _as_float_or_none(v):
    try:
        return float(v)
    except Exception:
        return None

def attach_dashboard_metadata(db, *,
                              targets_df=None,
                              formulas_df=None,
                              pillar_tests_df=None,
                              params_frames=None):
    """
    Build db.pillars and db.constants for the dashboard accordions.
    """
    pillars_meta = {}

    # pillar_tests -> role (ANCHOR, BENCHMARK, or VALIDATION)
    if pillar_tests_df is not None:
        pillar_col = None
        if "pillar_id" in pillar_tests_df.columns:
            pillar_col = "pillar_id"
        elif "label" in pillar_tests_df.columns:
            pillar_col = "label"

        if pillar_col and "role" in pillar_tests_df.columns:
            for _, r in pillar_tests_df.iterrows():
                pid = str(r.get(pillar_col, "")).strip()
                if not pid or pid.lower() == "nan":
                    continue
                role = str(r.get("role", "")).strip().upper()
                if role in ("ANCHOR", "BENCHMARK", "VALIDATION", "CALIBRATION"):
                    d = pillars_meta.setdefault(pid, {})
                    d["role"] = role

    # targets -> target, sigma, DOI
    # Check for both "pillar_id" and "label" columns
    if targets_df is not None:
        pillar_col = None
        if "pillar_id" in targets_df.columns:
            pillar_col = "pillar_id"
        elif "label" in targets_df.columns:
            pillar_col = "label"
        
        if pillar_col:
            for _, r in targets_df.iterrows():
                pid = str(r.get(pillar_col, "")).strip()
                if not pid:
                    continue
                d = pillars_meta.setdefault(pid, {})
                d["target"] = _as_float_or_none(r.get("target") or r.get("value") or r.get("target_value"))
                d["sigma"]  = _as_float_or_none(r.get("uncertainty") or r.get("sigma"))
                doi_raw = r.get("source_doi") or r.get("doi") or r.get("reference_doi")
                d["source_doi"] = _norm_doi(doi_raw)

    # formulas -> latex and name
    # Check for both "pillar_id" and "label" columns
    if formulas_df is not None:
        pillar_col = None
        if "pillar_id" in formulas_df.columns:
            pillar_col = "pillar_id"
        elif "label" in formulas_df.columns:
            pillar_col = "label"

        if pillar_col:
            for _, r in formulas_df.iterrows():
                pid = str(r.get(pillar_col, "")).strip()
                if not pid:
                    continue
                d = pillars_meta.setdefault(pid, {})
                d["latex"] = (r.get("latex") or r.get("equation_latex") or "").strip()
                # Also load name/description for formula display
                formula_name = r.get("name") or r.get("pillar_name") or r.get("description") or ""
                if formula_name:
                    d["name"] = str(formula_name).strip()

    # constants map from any param frames you have
    constants_map = {}
    for df in (params_frames or []):
        if df is None:
            continue
        for _, r in df.iterrows():
            raw = (r.get("input_token") or r.get("token") or r.get("name") or "")
            token = "" if raw is None else str(raw).strip()
            if token.lower() in ('nan', 'none', 'inf', '-inf'):
                token = ''
            if not token:
                continue
            val = r.get("value_si")
            constants_map[token.lower()] = _as_float_or_none(val) if _as_float_or_none(val) is not None else val

    # IMPORTANT: Don't overwrite existing pillars data from DBShim
    # Only add metadata that doesn't exist
    if hasattr(db, 'pillars') and db.pillars:
        for pid, meta in pillars_meta.items():
            if pid in db.pillars:
                db.pillars[pid].update({k: v for k, v in meta.items() if k not in ['target', 'sigma', 'unit']})
            else:
                db.pillars[pid] = meta
    else:
        setattr(db, "pillars", pillars_meta)
    
    setattr(db, "constants", constants_map)
    if not hasattr(db, "data_dir"):
        setattr(db, "data_dir", ".")

# local imports
ROOT = Path(__file__).parent.resolve()
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from UI.dashboard import DashboardGenerator  # renderer

# helpers

def _normpid(x):
    return re.sub(r"\s+", "", str(x).upper())

def _to_num(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in ["", "nan", "none"]:
        return None
    s = (
        s.replace("Ã—10^", "e")
         .replace("Ã—10", "e")
         .replace("E", "e")
         .replace("âˆ'", "-")
    )
    if "," in s and "." not in s:
        if s.count(",") == 1:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None

def _porder(pid):
    m = re.match(r"P(\d+)(B?)$", str(pid).strip().upper())
    if not m:
        return 9999
    n = int(m.group(1))
    return n + (0.5 if m.group(2) == "B" else 0.0)

# New tiny helper: never let parameter_type affect anything
def _strip_parameter_type(df):
    try:
        return df.drop(columns=["parameter_type"], errors="ignore")
    except Exception:
        return df


def _read_workbook(path: Path):
    """
    Loads all sheets and returns:
      targets, formulas, preds, pf, pc, po, pcoef, punits, ui_tooltips
    """
    try:
        print("[VERSION]", __RV_VERSION__)
        xl = pd.ExcelFile(path)
        print(f"[DEBUG] Successfully opened Excel file: {path}")

        targets  = xl.parse("Pillar_Targets", header=1)
        print(f"[DEBUG] Loaded Pillar_Targets: {targets.shape}")

        formulas = xl.parse("Pillar_Formulas", header=1)
        print(f"[DEBUG] Loaded Pillar_Formulas: {formulas.shape}")

        # Load Pillar_Tests for tooltip why/how content
        try:
            pillar_tests = xl.parse("Pillar_Tests", header=1)
            print(f"[DEBUG] Loaded Pillar_Tests: {pillar_tests.shape}")
        except Exception:
            pillar_tests = pd.DataFrame()
            print("[DEBUG] Pillar_Tests missing or unreadable; continuing")

        preds    = xl.parse("Model_Predictions_Matrix", header=1)
        print(f"[DEBUG] Loaded Model_Predictions_Matrix: {preds.shape}")

        # Parameters (drop optional parameter_type column defensively)
        pf     = _strip_parameter_type(xl.parse("Params_Fundamental", header=1))
        print(f"[DEBUG] Loaded Params_Fundamental: {pf.shape}")
        pc     = _strip_parameter_type(xl.parse("Params_Constants", header=1))
        print(f"[DEBUG] Loaded Params_Constants: {pc.shape}")
        po     = _strip_parameter_type(xl.parse("Params_Observational", header=1))
        print(f"[DEBUG] Loaded Params_Observational: {po.shape}")
        pcoef  = _strip_parameter_type(xl.parse("Params_Coefficients", header=1))
        print(f"[DEBUG] Loaded Params_Coefficients: {pcoef.shape}")
        punits = _strip_parameter_type(xl.parse("Params_Units", header=1))
        print(f"[DEBUG] Loaded Params_Units: {punits.shape}")

        # Tooltips are optional
        try:
            ui_tooltips = xl.parse("UI_Tooltips", header=1)
            print(f"[DEBUG] UI_Tooltips loaded: {ui_tooltips.shape[0]} rows, {ui_tooltips.shape[1]} columns")
            print(f"[DEBUG] UI_Tooltips columns: {list(ui_tooltips.columns)}")
        except Exception:
            ui_tooltips = pd.DataFrame()
            print("[DEBUG] UI_Tooltips missing or unreadable; continuing")

        # Model Registry for k values
        try:
            model_registry = xl.parse("Model_Registry", header=1)
            print(f"[DEBUG] Model_Registry loaded: {model_registry.shape}")
        except Exception:
            model_registry = pd.DataFrame()
            print("[DEBUG] Model_Registry missing or unreadable; continuing")

        print("[DEBUG] Successfully loaded all sheets, returning data...")
        return targets, formulas, pillar_tests, preds, pf, pc, po, pcoef, punits, ui_tooltips, model_registry
    except Exception as e:
        print("[ERROR] Failed reading workbook:", e)
        raise


def _pillars_1_to_13(targets, formulas, preds):
    def within_1_13(label):
        m = re.match(r"P(\d+)(B?)$", str(label).strip().upper())
        return bool(m and int(m.group(1)) <= 13)

    def is_vector_pillar(label):
        """Check if this is a vector pillar (P_SNE_PANTHEON, P_BAO_DESI, etc.)"""
        return str(label).strip().upper().startswith("P_")

    T = targets.copy()
    tid_col = "target_id" if "target_id" in T.columns else ("pillar_id" if "pillar_id" in T.columns else None)
    if tid_col is None:
        raise KeyError("Neither target_id nor pillar_id in Pillar_Targets")
    T["label"] = T[tid_col].astype(str).str.replace("target:", "", regex=False)
    T = T[T["label"].map(within_1_13)]

    F = formulas.copy()
    fid_col = "pillar_id" if "pillar_id" in F.columns else ("label" if "label" in F.columns else None)
    if fid_col is None:
        raise KeyError("Neither pillar_id nor label in Pillar_Formulas")
    F["label"] = F[fid_col]
    # Keep scalar pillars P1-P13 AND vector pillars (P_*)
    F = F[F["label"].map(lambda x: within_1_13(x) or is_vector_pillar(x))]

    P = preds.copy()
    if "pillar_id" in P.columns:
        P["label"] = P["pillar_id"]
        P = P[P["label"].map(within_1_13)]
    return T, F, P

def _build_token_store(pf, pc, po, pcoef, punits):
    base_map = {}
    unit_of = {}

    def add_df(df):
        for _, r in df.iterrows():
            v = r.get("value_si")
            token = r.get("input_token")
            unit = r.get("unit")
            if pd.notna(token) and str(token).strip():
                key = str(token).strip().lower().replace(" ", "").replace("_", "")
                base_map[key] = _to_num(v)
                unit_of[key] = str(unit or "").strip().lower()
            cid = r.get("canonical_id")
            if pd.notna(cid) and str(cid).strip():
                key = str(cid).strip().lower().replace(" ", "").replace("_", "")
                base_map[key] = _to_num(v)
                unit_of[key] = str(unit or "").strip().lower()

    for df in [pf, pc, po, pcoef, punits]:
        add_df(df)

    unit_conv = {}
    for _, r in punits.iterrows():
        token = str(r.get("input_token", "")).strip().lower()
        val = _to_num(r.get("value_si"))
        if token in ["kpc_to_m", "kiloparsec_to_m", "kiloparsec_to_meter"]:
            unit_conv["kpc"] = val
        if token in ["mpc_to_m", "mpc_to_meter", "megaparsec_to_meter"]:
            unit_conv["mpc"] = val
        if token in ["pc_to_m", "parsec_to_meter"]:
            unit_conv["pc"] = val

    for short in ["kpc", "mpc", "pc"]:
        if short in unit_conv:
            base_map[short] = unit_conv[short]
            unit_of[short] = "m"

    aliases = {
        "rotationscatterscaling": "Rotation Scatter Scaling",
        "environmentalfactor": "Environmental Rate Enhancement",
        "baselinerate": "Baseline Rate",
        "stressthreshold": "Stress Threshold",
        "avgstraineff": "Average Effective Strain",
        "avgstraineff10b": "Average Effective Strain (ultra-early band)",
        "ztypical": "Typical redshift z_typical",
        "zobs": "Observed redshift z_obs",
        "dcomoving": "Comoving path length d",
        "correlationfactor": "CMB TT Redshift Correlation Factor",
        "doverbeta10b": "Comoving-to-Beta Ratio for P10B",
        "kappa": "kappa_spec",
        "photoncouplingkappa": "kappa_spec",
        "photoncoupling": "kappa_spec",
        "angularcorrelation": "angular_corr_amp",
        "angularcorramp": "angular_corr_amp",
        "bases8": "Structure Growth S8 (baseline)",
    }

    # === DERIVED COEFFICIENTS FROM FUNDAMENTALS ===
    # These replace hand-tuned coefficients with formulas derived from {α, β_eos, z_t}
    def compute_derived_coefficients():
        alpha = base_map.get("alpha", 1.3)
        beta_eos = base_map.get("betaeos", base_map.get("beta_eos", 0.573))
        z_t = base_map.get("zt", base_map.get("z_t", 0.74))

        if alpha is None or beta_eos is None or z_t is None:
            print("[DERIVED] Warning: Missing fundamental parameters, using defaults")
            alpha = alpha or 1.3
            beta_eos = beta_eos or 0.573
            z_t = z_t or 0.74

        derived = {}

        # P3: base_correlation_adjusted = 1 - (1 - β_eos) / (1 + α)
        derived["derived_base_correlation"] = 1.0 - (1.0 - beta_eos) / (1.0 + alpha)

        # P6: stress_coupling = 2α / (1 + β_eos)
        derived["derived_stress_coupling"] = 2.0 * alpha / (1.0 + beta_eos)

        # P7: stress_threshold × environmental_factor = (α / β_eos) × 2(1 + β_eos)
        derived["derived_stress_env_product"] = (alpha / beta_eos) * 2.0 * (1.0 + beta_eos)

        # P11: cmb_tt_factor × z_typical = -β_eos/10 × z_t/(1+z_t)
        derived["derived_cmb_z_product"] = (-beta_eos / 10.0) * (z_t / (1.0 + z_t))

        # P13: c13_normalization = -(1 - β_eos) / (6α)
        derived["derived_c13"] = -(1.0 - beta_eos) / (6.0 * alpha)

        print(f"[DERIVED] Coefficients from fundamentals (α={alpha}, β_eos={beta_eos}, z_t={z_t}):")
        print(f"  P3  base_correlation = {derived['derived_base_correlation']:.6f}")
        print(f"  P6  stress_coupling  = {derived['derived_stress_coupling']:.6f}")
        print(f"  P7  stress×env       = {derived['derived_stress_env_product']:.6f}")
        print(f"  P11 cmb×z_product    = {derived['derived_cmb_z_product']:.6f}")
        print(f"  P13 c13_norm         = {derived['derived_c13']:.6f}")

        return derived

    derived_coeffs = compute_derived_coefficients()

    # Add derived coefficients to base_map (these override any hand-tuned values)
    for key, val in derived_coeffs.items():
        base_map[key.lower().replace("_", "")] = val

    def g(tok):
        key = tok.strip().lower().replace(" ", "").replace("_", "")
        if key in base_map and base_map[key] is not None:
            return base_map[key]

        if key in ("doverbeta", "d_over_beta"):
            d = g("d_comoving")
            b = g("beta")
            if isinstance(d, (int, float)) and isinstance(b, (int, float)) and b:
                return d / b
            return None
        if key in aliases:
            return g(aliases[key])
        return None

    # expose keys for quick debugging
    g._keys = sorted(set(base_map.keys()))
    g._derived = derived_coeffs
    return g

def _patch_formula_with_derived_coefficients(pillar_id, expr, inputs_csv):
    """
    Replace hand-tuned coefficients with derived expressions from fundamentals {α, β_eos, z_t}.
    Returns (patched_expr, patched_inputs).
    """
    original_expr = expr
    original_inputs = inputs_csv

    if pillar_id == "P3":
        # Replace base_correlation_adjusted with derived_base_correlation
        expr = expr.replace("base_correlation_adjusted", "derived_base_correlation")
        inputs_csv = inputs_csv.replace("base_correlation_adjusted", "derived_base_correlation")

    elif pillar_id == "P6":
        # Replace stress_coupling with derived_stress_coupling
        expr = expr.replace("stress_coupling", "derived_stress_coupling")
        inputs_csv = inputs_csv.replace("stress_coupling", "derived_stress_coupling")

    elif pillar_id == "P7":
        # Replace stress_threshold * environmental_factor with derived_stress_env_product
        expr = "(1 / 100.0) * (1 + derived_stress_env_product)"
        inputs_csv = "derived_stress_env_product"

    elif pillar_id == "P11":
        # Replace cmb_tt_correlation_factor * (z_typical/(1 + z_typical)) with derived_cmb_z_product
        expr = "1 - derived_cmb_z_product"
        inputs_csv = "derived_cmb_z_product"

    elif pillar_id == "P13":
        # Replace c13_normalization with derived_c13
        expr = expr.replace("c13_normalization", "derived_c13")
        inputs_csv = inputs_csv.replace("c13_normalization", "derived_c13")

    if expr != original_expr:
        print(f"[PATCH] {pillar_id}: '{original_expr}' → '{expr}'")

    return expr, inputs_csv


def _evaluate_mtdf(targets_df, formulas_df, get_token):
    # ---------- debug ----------
    print(f"[DEBUG] sparc_meas_uncertainty = {get_token('sparc_meas_uncertainty')}")
    print(f"[DEBUG] rar_deconv_factor = {get_token('rar_deconv_factor')}")

    # ---------- math env (NumPy ufuncs for array ops) ----------
    SAFE = {
        "sqrt": np.sqrt, "log": np.log, "log10": np.log10, "exp": np.exp,
        "sin": np.sin, "cos": np.cos, "tan": np.tan, "tanh": np.tanh,
        "pi": math.pi, "e": math.e, "abs": np.abs,
        "mean": lambda x: float(np.mean(np.asarray(x))),
        "clip": lambda x, lo, hi: np.clip(x, lo, hi),
        "rar_intrinsic_scatter": _rar_intrinsic_scatter,
    }
    print("[SAFE] numpy_ufuncs=", SAFE["sqrt"] is np.sqrt)

    # ---------- SPARC data ----------
    gobs, gbar = _load_sparc_profiles(_SPARC_PATH)
    print(f"[P1B] loaded arrays: g_obs={None if gobs is None else gobs.size}, g_bar={None if gbar is None else gbar.size}")

    # ---------- normalize pillar IDs ----------
    targets_df  = targets_df.copy()
    formulas_df = formulas_df.copy()
    targets_df["pid_norm"]  = targets_df["label"].map(_normpid)
    formulas_df["pid_norm"] = formulas_df["label"].map(_normpid)

    pillars      = list(targets_df["label"])
    pillars_norm = [_normpid(p) for p in pillars]

    # target map
    tgt_map = {}
    for _, r in targets_df.iterrows():
        tgt_map[r["pid_norm"]] = {"v": r.get("value"), "s": r.get("uncertainty"), "u": r.get("unit")}

    vals_norm = {}
    zs_norm   = {}

    # ---------- evaluate each formula ----------
    for _, r in formulas_df.iterrows():
        pidn = r["pid_norm"]
        if pidn not in tgt_map:
            continue

        label_for_log = r.get("pillar_id") or r.get("label") or "?"
        expr = str(r.get("python_expr", "") or "").replace("—", "-").replace("–", "-")
        raw_inputs = str(r.get("inputs_csv", "") or "")

        # Apply derived coefficient patches for P3, P6, P7, P11, P13
        expr, raw_inputs = _patch_formula_with_derived_coefficients(label_for_log, expr, raw_inputs)

        raw_inputs = str(raw_inputs or "")
        inputs = [t.strip() for t in re.split(r"[;,|]", raw_inputs) if t.strip()]
        local = {}

        # --- P1 prewire (before missing-check) ---
        if label_for_log == "P1":
            r_m_pw, v_obs_pw, v_bar_pw, v_err_pw = _load_sparc_velocity_fields(_SPARC_PATH)
            if r_m_pw is None:
                r_m_pw = _load_sparc_radii(_SPARC_PATH)
                if r_m_pw is not None and gobs is not None and gbar is not None:
                    v_obs_pw = np.sqrt(np.clip(gobs, 1e-30, None) * np.clip(r_m_pw[:gobs.size], 1e-9, None)) / 1e3
                    v_bar_pw = np.sqrt(np.clip(gbar, 1e-30, None) * np.clip(r_m_pw[:gbar.size], 1e-9, None)) / 1e3
                    v_err_pw = np.zeros_like(v_obs_pw)

            if r_m_pw is not None and v_obs_pw is not None and v_bar_pw is not None:
                local.update({
                    "r": r_m_pw,
                    "v_obs_kms": v_obs_pw,
                    "v_err_kms": v_err_pw if v_err_pw is not None else np.zeros_like(v_obs_pw),
                    "GM": (v_bar_pw * 1e3) ** 2 * r_m_pw,
                })
                # Do not ask token store for these; they are arrays
                inputs = [t for t in inputs if t not in ("GM", "r", "v_obs_kms", "v_err_kms")]
                print("[P1] prewire OK: r=", len(r_m_pw), "v_obs=", len(v_obs_pw))
            else:
                print("[P1] prewire: SPARC arrays unavailable")
        # --- end prewire ---

        # P1B wiring
        if label_for_log == "P1B":
            local.update({
                "g_obs_profile": gobs, "g_bar_profile": gbar,
                "g_obs": gobs, "g_bar": gbar,
                "distance": None, "inclination": None, "mass_to_light": None,
                "get_token": get_token,
            })
            inputs = []  # skip scalar get_token loop for P1B
        SAFE["rar_intrinsic_scatter"] = _rar_intrinsic_scatter

        # scalar tokens for all pillars (P1 will still need alpha/beta/ms_to_kms)
        missing = []
        for t in inputs:
            v = get_token(t)
            if v is None or (isinstance(v, float) and (np.isnan(v) or np.isinf(v))):
                missing.append(t)
            else:
                try:
                    local[t] = float(v)
                except Exception:
                    local[t] = v

        # extra fetch if expression references d_over_beta (your P10/P10B)
        if "d_over_beta" in expr and "d_over_beta" not in local:
            v = get_token("d_over_beta")
            if v is None:
                missing.append("d_over_beta")
            else:
                local["d_over_beta"] = float(v)

        if missing:
            print(f"[MISSING INPUTS] {label_for_log}: {', '.join(missing)}")
            vals_norm[pidn] = None
            continue

        # P10 / P10B debug
        if label_for_log in ("P10", "P10B"):
            def have(k): return k in local and local[k] is not None
            keys = ["kappa","d_over_beta","d_comoving","beta",
                    "avg_strain_eff","high_z_strain_scaler",
                    "high_z_d_over_beta_scaler","E","rho_cluster","c"]
            kv = [f"{k}={local.get(k) if have(k) else '<missing>'}" for k in keys]
            print(f"[CHECK {label_for_log}] " + ", ".join(kv))
            print(f"[EXPR {label_for_log}] {expr}")

        # P1 expr fallback if workbook expr is empty or truncated
        if label_for_log == "P1" and (not expr.strip() or "..." in expr):
            expr = (
                "sqrt(clip("
                "  mean((log10(v_obs_kms / (sqrt(GM / clip(r, 1e-9, 1e300) * (1 + alpha/(1 + r/beta))) * ms_to_kms)))**2)"
                "  - mean((v_err_kms / (2.302585093 * clip(v_obs_kms, 1e-300, 1e300)))**2),"
                "  0.0, 1e300))"
            )

        # evaluate
        try:
            val = eval(expr, {"__builtins__": {}}, {**SAFE, **local})

            conv_expr = str(r.get("target_conversion_expr", "") or "")
            conv_flag = conv_expr.strip().lower()
            if conv_flag not in ("", "result", "direct"):
                try:
                    val = eval(conv_expr, {"__builtins__": {}}, {**SAFE, **local, "result": val})
                except Exception as e:
                    print(f"[WARN] target conversion failed for {label_for_log}: {e}  expr='{conv_expr}'")

            vals_norm[pidn] = float(val)

        except Exception as e:
            print(f"[ERROR] Eval failed for {label_for_log}: {e}  expr='{expr}'")
            # P1 dedicated fallback (compute directly with NumPy)
            if label_for_log == "P1":
                try:
                    alpha     = local.get("alpha")     if "alpha"     in local else get_token("alpha")
                    beta      = local.get("beta")      if "beta"      in local else get_token("beta")
                    ms_to_kms = local.get("ms_to_kms") if "ms_to_kms" in local else get_token("ms_to_kms")

                    r_arr  = np.asarray(local["r"])
                    v_obs  = np.asarray(local["v_obs_kms"])
                    v_err  = np.asarray(local.get("v_err_kms", np.zeros_like(v_obs)))
                    GM_arr = np.asarray(local["GM"])

                    denom = np.sqrt(GM_arr / np.clip(r_arr, 1e-9, None) * (1.0 + alpha / (1.0 + r_arr / beta))) * ms_to_kms
                    ratio = v_obs / denom
                    core  = np.log10(np.clip(ratio, 1e-300, None))
                    meas  = v_err / (2.302585093 * np.clip(v_obs, 1e-300, None))
                    vals_norm[pidn] = float(np.sqrt(np.clip(np.mean(core**2) - np.mean(meas**2), 0.0, 1e300)))
                    print("[P1] fallback_eval=", vals_norm[pidn])
                except Exception as e2:
                    print("[P1] fallback_eval failed:", e2)
                    vals_norm[pidn] = None
            else:
                vals_norm[pidn] = None

    # --- P1 post-processing fallback ---
    try:
        if "P1" in pillars_norm and vals_norm.get("P1") is None:
            print("[P1] post_fallback: computing directly (loop did not fill P1)")
            r_m_pf, v_obs_pf, v_bar_pf, v_err_pf = _load_sparc_velocity_fields(_SPARC_PATH)
            if r_m_pf is None:
                r_m_pf = _load_sparc_radii(_SPARC_PATH)
                if r_m_pf is not None and gobs is not None and gbar is not None:
                    v_obs_pf = np.sqrt(np.clip(gobs, 1e-30, None) * np.clip(r_m_pf[:gobs.size], 1e-9, None)) / 1e3
                    v_bar_pf = np.sqrt(np.clip(gbar, 1e-30, None) * np.clip(r_m_pf[:gbar.size], 1e-9, None)) / 1e3
                    v_err_pf = np.zeros_like(v_obs_pf)
            if r_m_pf is not None and v_obs_pf is not None and v_bar_pf is not None:
                alpha     = get_token("alpha")
                beta      = get_token("beta")
                ms_to_kms = get_token("ms_to_kms")
                GM_pf = (v_bar_pf * 1e3) ** 2 * r_m_pf
                denom = np.sqrt(GM_pf / np.clip(r_m_pf, 1e-9, None) * (1.0 + alpha / (1.0 + r_m_pf / beta))) * ms_to_kms
                ratio = v_obs_pf / denom
                core  = np.log10(np.clip(ratio, 1e-300, None))
                meas  = (v_err_pf if v_err_pf is not None else np.zeros_like(v_obs_pf)) / (2.302585093 * np.clip(v_obs_pf, 1e-300, None))
                vals_norm["P1"] = float(np.sqrt(np.clip(np.mean(core**2) - np.mean(meas**2), 0.0, 1e300)))
                print("[P1] post_fallback OK:", vals_norm["P1"])
            else:
                print("[P1] post_fallback: SPARC arrays not available")
    except Exception as e:
        print("[P1] post_fallback error:", e)
    # --- end P1 post-processing fallback ---

    # ---------- z-scores / chi2 ----------
    chi2   = 0.0
    n      = 0
    passes = 0
    for pidn in pillars_norm:
        v  = vals_norm.get(pidn)
        tgt = tgt_map.get(pidn, {})
        tv, ts = tgt.get("v"), tgt.get("s")
        if v is None or tv is None or ts in (None, 0) or pd.isna(tv) or pd.isna(ts):
            continue
        z = (v - tv) / ts
        # Guard against NaN/inf z-scores corrupting chi2
        if np.isnan(z) or np.isinf(z):
            print(f"[WARN] Skipping pillar {pidn}: z-score is {z} (v={v}, tv={tv}, ts={ts})")
            continue
        zs_norm[pidn] = float(z)
        chi2 += float(z * z)
        n += 1
        if abs(z) <= 1.0:
            passes += 1

    dof = n
    chi2_red = (chi2 / dof) if dof else float("nan")
    return pillars, pillars_norm, tgt_map, vals_norm, zs_norm, passes, dof, chi2, chi2_red

    
def _build_k_map(model_registry):
    """
    Build a map of model short names to k_total (free parameter count).
    Returns dict like {"MTDF": 0, "LCDM": 6, "MOND": 1, ...}
    Also includes variations like "LCDM_PRED_VALUE" to match predictions matrix columns.
    """
    k_map = {}
    if model_registry is None or model_registry.empty:
        print("[DEBUG] Model_Registry empty, using defaults for k values")
        # Fallback defaults
        return {
            "MTDF": 0,
            "LCDM": 6,
            "MOND": 1,
            "EDE": 9,
            "FDM": 7,
            "SIDM": 7,
        }

    # Try to find short name and k_total columns (with fallback to k)
    for _, row in model_registry.iterrows():
        short = row.get("short") or row.get("model_short") or row.get("model_id")
        k = row.get("k_total") or row.get("k")  # Prefer k_total, fallback to k
        if pd.notna(short) and pd.notna(k):
            try:
                # Normalize to ASCII and uppercase
                short_ascii = str(short).replace("Λ", "L").replace("λ", "l").strip().upper()
                k_val = int(k)

                # Add main variant
                k_map[short_ascii] = k_val

                # Add common variations (for matching predictions matrix columns)
                k_map[f"{short_ascii}_PRED_VALUE"] = k_val
                k_map[f"{short_ascii}_PRED_UNCERTAINTY"] = k_val

                print(f"[K_MAP] {short_ascii} → k_total={k_val} (+ variants)")
            except Exception:
                pass

    # If nothing was loaded, fall back to defaults
    if not k_map:
        print("[DEBUG] No k values loaded from Model_Registry, using fallback defaults")
        return {
            "MTDF": 0,
            "LCDM": 6,
            "MOND": 1,
            "EDE": 9,
            "FDM": 7,
            "SIDM": 7,
        }

    return k_map


def _mk_row_from_vals(name, pillars, pillars_norm, tgt_map, vals_map_norm,
                      foundation="PUBLISHED", proof_sigma=1.0, k=0):
    zs_local = {}
    chi2_local = 0.0
    tested_local = 0
    for pidn in pillars_norm:
        v = vals_map_norm.get(pidn)
        tv = tgt_map.get(pidn, {}).get("v")
        ts = tgt_map.get(pidn, {}).get("s")
        if v is None or tv is None or ts in (None, 0) or pd.isna(tv) or pd.isna(ts):
            continue
        z = (v - tv) / ts
        zs_local[pidn] = z
        chi2_local += z * z
        tested_local += 1

    dof_local = tested_local
    chi2_red_local = (chi2_local / dof_local) if dof_local else float("nan")

    vals = {label: vals_map_norm.get(pidn) for label, pidn in zip(pillars, pillars_norm)}
    zs = {label: zs_local.get(pidn) for label, pidn in zip(pillars, pillars_norm)}
    maxz = max((abs(z) for z in zs.values() if z is not None), default=0.0)

    passes_ct = sum(1 for z in zs.values() if z is not None and abs(z) <= proof_sigma)
    proof_frac = (passes_ct / dof_local) if dof_local else 0.0
    proof_pct  = 100.0 * proof_frac

    return {
        "model": name,
        "foundation_type": foundation,
        "vals": vals,
        "z_scores": zs,
        "passes": passes_ct,
        "dof": dof_local,
        "proof": proof_frac,
        "proof_frac": proof_frac,
        "proof_percent": proof_pct,
        "proof_pct": proof_pct,
        "chi2": chi2_local,
        "chi2_red": chi2_red_local,
        "max_z": maxz,
        "AIC": float("nan"),
        "BIC": float("nan"),
        "k": k,
        "evidence": 0,
        "tier_code": "VALIDATED" if foundation.upper() == "EMPIRICAL" else "REQUIRES_COMPONENTS",
    }

def _extract_model_rows_from_matrix(preds_df, pillars, pillars_norm, tgt_map, k_map=None, proof_sigma=1.0):
    rows = []
    if preds_df.empty:
        return rows

    if k_map is None:
        k_map = {}

    long_cols = {c.lower().strip() for c in preds_df.columns}
    if {"pillar_id", "model", "prediction"}.issubset(long_cols):
        df = preds_df.rename(columns={c: c.lower().strip() for c in preds_df.columns})
        for model, g in df.groupby("model"):
            vals_norm = {}
            for _, r in g.iterrows():
                pidn = _normpid(r["pillar_id"])
                vals_norm[pidn] = _to_num(r["prediction"])
            model_upper = str(model).strip().upper()
            k = k_map.get(model_upper, 0)
            rows.append(_mk_row_from_vals(model, pillars, pillars_norm, tgt_map,
                                          vals_norm, foundation="PUBLISHED",
                                          proof_sigma=proof_sigma, k=k))
        return rows

    meta_cols = {
        "pillar_id","label","unit","notes","updated_at","updated_by","canonical_id",
        "class","output_id","latex","python_expr","target_conversion_expr",
        "target_conversion_reason","inputs_csv","dependencies","reference_doi",
        "description","Disp_citation_ids","pillar_scope",
    }
    candidates = [c for c in preds_df.columns if str(c) not in meta_cols]
    candidates = [c for c in candidates if c not in ["index"]]

    model_cols = []
    for c in candidates:
        series = preds_df[c]
        has_num = pd.to_numeric(series, errors="coerce").notna().any()
        if has_num and c not in ["pillar_id", "label"]:
            model_cols.append(c)

    for c in model_cols:
        vals_norm = {}
        for _, r in preds_df.iterrows():
            pidn = _normpid(r.get("label") or r.get("pillar_id"))
            v = _to_num(r.get(c))
            if pidn and v is not None:
                vals_norm[pidn] = v
        model_upper = str(c).strip().upper()
        k = k_map.get(model_upper, 0)
        rows.append(_mk_row_from_vals(str(c), pillars, pillars_norm, tgt_map,
                                      vals_norm, foundation="PUBLISHED",
                                      proof_sigma=proof_sigma, k=k))
    return rows

def main():
    ap = argparse.ArgumentParser(description="Automated multi-model dashboard with MTDF engine")
    ap.add_argument("--workbook", required=True, help="Path to DB_Workbook_STRICT_*.xlsx")
    ap.add_argument("--out", required=True, help="Output HTML path")
    ap.add_argument("--diag", default="Diagnostics.csv", help="Diagnostics CSV path")
    ap.add_argument("--proof_sigma", type=float, default=1.0, help="Threshold in sigma for Proof")
    args = ap.parse_args()

    wb = Path(args.workbook).resolve()
    if not wb.exists():
        raise FileNotFoundError(f"Workbook not found: {wb}")

    targets, formulas, pillar_tests, preds, pf, pc, po, pcoef, punits, ui_tooltips, model_registry = _read_workbook(wb)
    targets, formulas, preds = _pillars_1_to_13(targets, formulas, preds)

    # Build k_map from Model_Registry
    k_map = _build_k_map(model_registry)
    print(f"[K_MAP] Loaded {len(k_map)} model k values")

    # MTDF evaluation (scalar pillars)
    get_token = _build_token_store(pf, pc, po, pcoef, punits)
    print(f"[TOKENS] available: {len(get_token._keys)}")
    pillars, pillars_norm, tgt_map, mtdf_vals, mtdf_zs, mtdf_passes, mtdf_dof, mtdf_chi2, mtdf_chi2_red = _evaluate_mtdf(
        targets, formulas, get_token
    )

    # Vector pillar evaluation (SNe, BAO, H(z), fσ₈, CMB) - workbook-driven
    # Baseline MTDF (no early field energy)
    vector_stats = _evaluate_vector_pillars(get_token, data_dir=str(wb.parent), workbook_path=str(wb), use_efe_cmb=False)

    # MTDF (EFE) - with Early Field Energy correction for CMB
    vector_stats_efe = _evaluate_vector_pillars(get_token, data_dir=str(wb.parent), workbook_path=str(wb), use_efe_cmb=True)

    # Combined MTDF statistics (scalar + vector)
    combined_chi2 = mtdf_chi2 + vector_stats["total_chi2"]
    combined_dof = mtdf_dof + vector_stats["total_dof"]
    combined_chi2_red = combined_chi2 / combined_dof if combined_dof > 0 else float("nan")

    print(f"[COMBINED] Scalar DOF: {mtdf_dof}, Vector DOF: {vector_stats['total_dof']}, Total DOF: {combined_dof}")
    print(f"[COMBINED] Scalar χ²: {mtdf_chi2:.2f}, Vector χ²: {vector_stats['total_chi2']:.2f}, Total χ²: {combined_chi2:.2f}")
    print(f"[COMBINED] χ²/ν = {combined_chi2_red:.4f}")

    # Combined MTDF statistics EXCLUDING CMB (to show late-time fit quality)
    cmb_chi2 = 0.0
    cmb_dof = 0
    for vr in vector_stats.get("vector_results", []):
        if "CMB" in vr.get("pillar_id", "").upper() or "CMB" in vr.get("name", "").upper():
            cmb_chi2 = vr.get("chi2", 0.0)
            cmb_dof = vr.get("dof", 0)
            break

    combined_excl_cmb_chi2 = combined_chi2 - cmb_chi2
    combined_excl_cmb_dof = combined_dof - cmb_dof
    combined_excl_cmb_chi2_red = combined_excl_cmb_chi2 / combined_excl_cmb_dof if combined_excl_cmb_dof > 0 else float("nan")

    print(f"[COMBINED excl. CMB] χ²: {combined_excl_cmb_chi2:.2f}, DOF: {combined_excl_cmb_dof}, χ²/ν = {combined_excl_cmb_chi2_red:.4f}")

    mtdf_k = k_map.get("MTDF", 0)
    mtdf_row = _mk_row_from_vals("MTDF", pillars, pillars_norm, tgt_map, mtdf_vals,
                                 foundation="EMPIRICAL", proof_sigma=args.proof_sigma, k=mtdf_k)

    # Add vector pillar stats to MTDF row
    mtdf_row["vector_chi2"] = vector_stats["total_chi2"]
    mtdf_row["vector_dof"] = vector_stats["total_dof"]
    mtdf_row["combined_chi2"] = combined_chi2
    mtdf_row["combined_dof"] = combined_dof
    mtdf_row["combined_chi2_red"] = combined_chi2_red
    mtdf_row["vector_results"] = vector_stats["vector_results"]
    # Excluding CMB stats (for late-time fit display)
    mtdf_row["combined_excl_cmb_chi2"] = combined_excl_cmb_chi2
    mtdf_row["combined_excl_cmb_dof"] = combined_excl_cmb_dof
    mtdf_row["combined_excl_cmb_chi2_red"] = combined_excl_cmb_chi2_red
    mtdf_row["cmb_chi2"] = cmb_chi2
    mtdf_row["cmb_dof"] = cmb_dof

    # Build extended pillar list including vector pillars
    vector_pillar_ids = []
    vector_pillar_meta = {}
    for vr in vector_stats.get("vector_results", []):
        pid = vr.get("pillar_id", "")
        if pid:
            vector_pillar_ids.append(pid)
            # Add vector pillar value to MTDF vals (use χ²/ν as the "value")
            mtdf_row["vals"][pid] = vr.get("chi2_red", float("nan"))
            # Store chi2 and dof for tooltip display
            mtdf_row.setdefault("vector_pillar_data", {})[pid] = {
                "chi2": vr.get("chi2", 0),
                "dof": vr.get("dof", 0),
                "chi2_red": vr.get("chi2_red", float("nan")),
                "n_data": vr.get("n_data", 0),
            }
            # Metadata for DB shim
            vector_pillar_meta[pid] = {
                "name": vr.get("name", pid),
                "category": vr.get("category", "VECTOR"),
                "pillar_mode": "VECTOR",
                "n_data": vr.get("n_data", 0),
                "dof": vr.get("dof", 0),
            }

    # Update MTDF passes and DOF to include vector pillars
    # Vector pillar passes if χ²/ν < 1.5 (green threshold)
    scalar_passes = mtdf_row.get("passes", 0)
    scalar_dof = mtdf_row.get("dof", 0)
    scalar_chi2 = mtdf_row.get("chi2", 0.0)
    scalar_chi2_red = mtdf_row.get("chi2_red", float("nan"))

    # Store original scalar values before overwriting with combined
    mtdf_row["scalar_dof"] = scalar_dof
    mtdf_row["scalar_chi2"] = scalar_chi2
    mtdf_row["scalar_chi2_red"] = scalar_chi2_red
    vector_passes = 0
    vector_count = 0
    for vr in vector_stats.get("vector_results", []):
        if not vr.get("experimental", False):  # Only count non-experimental
            vector_count += 1
            chi2_red = vr.get("chi2_red", float("inf"))
            if chi2_red < 1.5:  # Green threshold for vector pillars
                vector_passes += 1

    # Update MTDF row with combined passes and DOF
    mtdf_row["passes"] = scalar_passes + vector_passes
    mtdf_row["dof"] = scalar_dof + vector_count
    # Recalculate proof fraction
    if mtdf_row["dof"] > 0:
        mtdf_row["proof"] = mtdf_row["passes"] / mtdf_row["dof"]
        mtdf_row["proof_frac"] = mtdf_row["proof"]
        mtdf_row["proof_pct"] = 100.0 * mtdf_row["proof"]
    # Strict vs diagnostic summary (CMB* excluded from strict totals)
    cmb_pid = "P_CMB_DIST"
    cmb_is_counted = False
    cmb_chi2_red_here = float("nan")
    cmb_pass_here = 0
    strict_vector_count = 0
    strict_vector_passes = 0
    for vr in vector_stats.get("vector_results", []):
        if vr.get("experimental", False):
            continue
        pid = vr.get("pillar_id", "") or ""
        is_cmb = (pid == cmb_pid) or ("CMB" in pid.upper()) or ("CMB" in (vr.get("name", "") or "").upper())
        if is_cmb:
            cmb_is_counted = True
            cmb_chi2_red_here = vr.get("chi2_red", float("inf"))
            cmb_pass_here = 1 if cmb_chi2_red_here < 1.5 else 0
            continue
        strict_vector_count += 1
        if (vr.get("chi2_red", float("inf")) < 1.5):
            strict_vector_passes += 1

    strict_passes = scalar_passes + strict_vector_passes
    strict_tests = scalar_dof + strict_vector_count
    strict_pct = (100.0 * strict_passes / strict_tests) if strict_tests > 0 else float("nan")

    # Prefer the precomputed strict χ²/ν excluding CMB if available
    strict_chi2_red = mtdf_row.get("combined_excl_cmb_chi2_red", float("nan"))
    combined_chi2_red_all = mtdf_row.get("combined_chi2_red", mtdf_row.get("chi2_red", float("nan")))

    print(f"[PASSES] Strict (excl CMB*): {strict_passes}/{strict_tests} ({strict_pct:.1f}%)")
    if cmb_is_counted:
        print(f"[DIAGNOSTIC] CMB* χ²/ν: {cmb_chi2_red_here:.4f} (excluded from strict totals)")
    print(f"[TOTALS] Combined χ²/ν (incl diagnostics): {combined_chi2_red_all:.4f}")
    print(f"[TOTALS] Strict χ²/ν (fit pillars only): {strict_chi2_red:.4f}")

    # =========================================================================
    # MTDF (EFE) Row - MTDF with Early Field Energy correction for CMB
    # =========================================================================
    # Create a duplicate of MTDF row but with EFE vector stats (only CMB differs)
    mtdf_efe_row = _mk_row_from_vals("MTDF (EFE)", pillars, pillars_norm, tgt_map, mtdf_vals,
                                      foundation="EMPIRICAL", proof_sigma=args.proof_sigma, k=mtdf_k)

    # Combined EFE statistics (scalar + vector with EFE CMB)
    combined_chi2_efe = mtdf_chi2 + vector_stats_efe["total_chi2"]
    combined_dof_efe = mtdf_dof + vector_stats_efe["total_dof"]
    combined_chi2_red_efe = combined_chi2_efe / combined_dof_efe if combined_dof_efe > 0 else float("nan")

    # EFE CMB stats
    cmb_chi2_efe = 0.0
    cmb_dof_efe = 0
    for vr in vector_stats_efe.get("vector_results", []):
        if "CMB" in vr.get("pillar_id", "").upper() or "CMB" in vr.get("name", "").upper():
            cmb_chi2_efe = vr.get("chi2", 0.0)
            cmb_dof_efe = vr.get("dof", 0)
            break

    combined_excl_cmb_chi2_efe = combined_chi2_efe - cmb_chi2_efe
    combined_excl_cmb_dof_efe = combined_dof_efe - cmb_dof_efe
    combined_excl_cmb_chi2_red_efe = combined_excl_cmb_chi2_efe / combined_excl_cmb_dof_efe if combined_excl_cmb_dof_efe > 0 else float("nan")

    if cmb_chi2_efe > 0:
        print(f"[MTDF (EFE)] CMB χ²: {cmb_chi2_efe:.2f} (baseline was {cmb_chi2:.2f}, improvement factor: {cmb_chi2/cmb_chi2_efe:.1f}x)")
    else:
        print(f"[MTDF (EFE)] CMB χ²: {cmb_chi2_efe:.2f} (baseline was {cmb_chi2:.2f}, no CMB data available)")

    # Add vector pillar stats to EFE row
    mtdf_efe_row["vector_chi2"] = vector_stats_efe["total_chi2"]
    mtdf_efe_row["vector_dof"] = vector_stats_efe["total_dof"]
    mtdf_efe_row["combined_chi2"] = combined_chi2_efe
    mtdf_efe_row["combined_dof"] = combined_dof_efe
    mtdf_efe_row["combined_chi2_red"] = combined_chi2_red_efe
    mtdf_efe_row["vector_results"] = vector_stats_efe["vector_results"]
    mtdf_efe_row["combined_excl_cmb_chi2"] = combined_excl_cmb_chi2_efe
    mtdf_efe_row["combined_excl_cmb_dof"] = combined_excl_cmb_dof_efe
    mtdf_efe_row["combined_excl_cmb_chi2_red"] = combined_excl_cmb_chi2_red_efe
    mtdf_efe_row["cmb_chi2"] = cmb_chi2_efe
    mtdf_efe_row["cmb_dof"] = cmb_dof_efe
    mtdf_efe_row["is_efe"] = True  # Flag to identify this as EFE row

    # Add vector pillar values to EFE row
    for vr in vector_stats_efe.get("vector_results", []):
        pid = vr.get("pillar_id", "")
        if pid:
            mtdf_efe_row["vals"][pid] = vr.get("chi2_red", float("nan"))
            mtdf_efe_row.setdefault("vector_pillar_data", {})[pid] = {
                "chi2": vr.get("chi2", 0),
                "dof": vr.get("dof", 0),
                "chi2_red": vr.get("chi2_red", float("nan")),
                "n_data": vr.get("n_data", 0),
                "efe_mode": vr.get("efe_mode", False),
            }

    # Store scalar stats for EFE row
    mtdf_efe_row["scalar_dof"] = scalar_dof
    mtdf_efe_row["scalar_chi2"] = scalar_chi2
    mtdf_efe_row["scalar_chi2_red"] = scalar_chi2_red

    # Calculate passes for EFE row
    efe_vector_passes = 0
    efe_vector_count = 0
    for vr in vector_stats_efe.get("vector_results", []):
        if not vr.get("experimental", False):
            efe_vector_count += 1
            chi2_red = vr.get("chi2_red", float("inf"))
            if chi2_red < 1.5:
                efe_vector_passes += 1

    mtdf_efe_row["passes"] = scalar_passes + efe_vector_passes
    mtdf_efe_row["dof"] = scalar_dof + efe_vector_count
    if mtdf_efe_row["dof"] > 0:
        mtdf_efe_row["proof"] = mtdf_efe_row["passes"] / mtdf_efe_row["dof"]
        mtdf_efe_row["proof_frac"] = mtdf_efe_row["proof"]
        mtdf_efe_row["proof_pct"] = 100.0 * mtdf_efe_row["proof"]

    # Strict vs diagnostic summary for MTDF (EFE)
    cmb_pid = "P_CMB_DIST"
    cmb_is_counted = False
    cmb_chi2_red_here = float("nan")
    strict_vector_count = 0
    strict_vector_passes = 0
    for vr in vector_stats_efe.get("vector_results", []):
        if vr.get("experimental", False):
            continue
        pid = vr.get("pillar_id", "") or ""
        is_cmb = (pid == cmb_pid) or ("CMB" in pid.upper()) or ("CMB" in (vr.get("name", "") or "").upper())
        if is_cmb:
            cmb_is_counted = True
            cmb_chi2_red_here = vr.get("chi2_red", float("inf"))
            continue
        strict_vector_count += 1
        if (vr.get("chi2_red", float("inf")) < 1.5):
            strict_vector_passes += 1

    strict_passes_efe = scalar_passes + strict_vector_passes
    strict_tests_efe = scalar_dof + strict_vector_count
    strict_pct_efe = (100.0 * strict_passes_efe / strict_tests_efe) if strict_tests_efe > 0 else float("nan")

    strict_chi2_red_efe = mtdf_efe_row.get("combined_excl_cmb_chi2_red", float("nan"))
    combined_chi2_red_all_efe = mtdf_efe_row.get("combined_chi2_red", mtdf_efe_row.get("chi2_red", float("nan")))

    print(f"[PASSES EFE] Strict (excl CMB*): {strict_passes_efe}/{strict_tests_efe} ({strict_pct_efe:.1f}%)")
    if cmb_is_counted:
        print(f"[DIAGNOSTIC EFE] CMB* χ²/ν: {cmb_chi2_red_here:.4f} (excluded from strict totals)")
    print(f"[TOTALS EFE] Combined χ²/ν (incl diagnostics): {combined_chi2_red_all_efe:.4f}")
    print(f"[TOTALS EFE] Strict χ²/ν (fit pillars only): {strict_chi2_red_efe:.4f}")


    # =========================================================================

    # Combined pillar list: scalar P1-P13 + vector pillars
    all_pillars = pillars + vector_pillar_ids

    # External models from predictions matrix (scalar pillars only)
    other_rows = _extract_model_rows_from_matrix(preds, pillars, pillars_norm, tgt_map,
                                                 k_map=k_map, proof_sigma=args.proof_sigma)

    # Load literature values for vector pillars from workbook
    literature_values = {}
    try:
        lit_df = pd.read_excel(wb, sheet_name='Vector_Pillar_Literature', header=3)
        for _, lit_row in lit_df.iterrows():
            pillar_id = lit_row.get('pillar_id')
            model = str(lit_row.get('model', '')).upper()
            chi2_nu = lit_row.get('chi2_nu_lit')
            chi2 = lit_row.get('chi2_lit')
            dof = lit_row.get('dof_lit')
            ref = lit_row.get('reference', '')
            notes = lit_row.get('notes', '')
            if pillar_id and model:
                key = (pillar_id, model)
                literature_values[key] = {
                    'chi2_nu': chi2_nu if pd.notna(chi2_nu) else None,
                    'chi2': chi2 if pd.notna(chi2) else None,
                    'dof': int(dof) if pd.notna(dof) else None,
                    'reference': ref,
                    'notes': notes,
                    'is_literature': True
                }
        print(f"[LITERATURE] Loaded {len(literature_values)} vector pillar literature values")
    except Exception as e:
        print(f"[LITERATURE] No literature values loaded: {e}")

    # Add literature values or n/a for vector pillars in other models
    for row in other_rows:
        model_name = row.get('model', '').upper()
        for vpid in vector_pillar_ids:
            lit_key = (vpid, model_name)
            if lit_key in literature_values:
                lit_data = literature_values[lit_key]
                # Use chi2_nu as the displayed value (same as MTDF)
                row["vals"][vpid] = lit_data['chi2_nu']
                # Store literature metadata for tooltip
                row.setdefault("vector_pillar_data", {})[vpid] = {
                    "chi2": lit_data['chi2'],
                    "dof": lit_data['dof'],
                    "chi2_red": lit_data['chi2_nu'],
                    "n_data": lit_data['dof'],  # Approximate
                    "is_literature": True,
                    "reference": lit_data['reference'],
                    "notes": lit_data['notes'],
                }
            else:
                row["vals"][vpid] = None  # n/a - no literature value found

    rows = [mtdf_row, mtdf_efe_row] + other_rows

    # DB shim for UI
    class DBShim:
        def __init__(self):
            # Scalar pillars from targets
            self.pillars = {
                r["label"]: {
                    "target": r["value"],
                    "sigma": r["uncertainty"],
                    "unit": r["unit"],
                    "name": r.get("pillar_name", r["label"]),
                    "pillar_mode": "SCALAR",
                }
                for _, r in targets.iterrows()
            }
            # Add vector pillars
            for vpid, vmeta in vector_pillar_meta.items():
                self.pillars[vpid] = {
                    "target": None,  # Vector pillars don't have single target
                    "sigma": None,
                    "unit": "χ²/ν",
                    "name": vmeta["name"],
                    "pillar_mode": "VECTOR",
                    "n_data": vmeta["n_data"],
                    "dof": vmeta["dof"],
                    "category": vmeta["category"],
                }
            self.models = {}
            self.data_dir = str(wb.parent)

    db = DBShim()

    # Attach metadata for bottom panels
    tables = {
        "Pillar_Targets": targets,
        "Pillar_Formulas": formulas,
        "Params_Constants": pc,
        "Params_Fundamentals": pf,
        "Params_Observational": po,
        "Params_Coefficients": pcoef,
    }
    def pick(*keys):
        for k in keys:
            if k in tables and tables[k] is not None:
                return tables[k]
        return None

    df_pillar_targets = pick("Pillar_Targets", "DB_Pillar_Targets", "Targets")
    df_pillar_formulas = pick("Pillar_Formulas", "DB_Pillar_Formulas", "Formulas")
    df_params_constants = pick("Params_Constants", "Constants", "DB_Constants")
    df_params_fundamentals = pick("Params_Fundamentals", "Fundamentals", "DB_Fundamentals")
    df_params_observational = pick("Params_Observational", "Observational", "DB_Observational")
    df_params_coefficients = pick("Coefficients", "Params_Coefficients", "DB_Coefficients")

    attach_dashboard_metadata(
        db,
        targets_df=df_pillar_targets,
        formulas_df=df_pillar_formulas,
        pillar_tests_df=pillar_tests,
        params_frames=[
            df_params_constants,
            df_params_fundamentals,
            df_params_observational,
            df_params_coefficients,
        ],
    )

    # Create tooltip engine with workbook data
    try:
        from UI.tooltip_engine import TooltipEngine
        workbook_data = {
            'Pillar_Targets': targets,
            'Pillar_Formulas': formulas,
            'Pillar_Tests': pillar_tests,  # Actual Pillar_Tests sheet with why/how content
            'Params_Constants': pc,
            'Params_Fundamental': pf,
            'Params_Observational': po,
            'Params_Coefficients': pcoef,
            'Params_Units': punits,
            'UI_Tooltips': ui_tooltips,
        }
        tooltip_engine = TooltipEngine(workbook_data)
        print(f"[TOOLTIPS] Loaded {len(tooltip_engine.get_all_tooltip_ids())} tooltip definitions")
    except ImportError:
        print("[WARNING] Tooltip engine not available - using fallback")
        tooltip_engine = None

    # Render HTML with new tooltip engine
    gen = DashboardGenerator(db=db, tooltip_engine=tooltip_engine)

    # Render HTML
    html = gen.generate_complete_dashboard({"rows": rows, "pillars": all_pillars})
    Path(args.out).write_text(html, encoding="utf-8")

    # Diagnostics CSV for MTDF
    diag_rows = []
    for label, pidn in zip(pillars, pillars_norm):
        v = mtdf_vals.get(pidn)
        tgt = tgt_map.get(pidn, {})
        tv, ts = tgt.get("v"), tgt.get("s")
        z = mtdf_zs.get(pidn)
        if v is None or tv is None or ts in (None, 0) or z is None:
            continue
        contrib = float(z * z)
        diag_rows.append({
            "pillar": label,
            "prediction": v,
            "target": tv,
            "sigma": ts,
            "z": z,
            "chi2_contrib": contrib,
            "unit": tgt.get("u", "")
        })
    diag_df = pd.DataFrame(diag_rows).sort_values("chi2_contrib", ascending=False)
    Path(args.diag).write_text(diag_df.to_csv(index=False), encoding="utf-8")

    # Console summary
    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("=" * 60)
    print("Run complete:", now)
    print("Workbook:", wb)
    print("HTML out:", Path(args.out).resolve())
    print("Diagnostics CSV (MTDF):", Path(args.diag).resolve())
    print("Models included:", ", ".join([r["model"] for r in rows]))
    print("-" * 60)
    print("MTDF SCALAR PILLARS:")
    print(f"  Tested pillars: {mtdf_row['dof']}")
    print(f"  χ² = {mtdf_row['chi2']:.3f}, χ²/ν = {mtdf_row['chi2_red']:.4f}")
    print(f"  Proof: {mtdf_row['passes']}/{mtdf_row['dof']} = {mtdf_row['proof_pct']:.1f}% at {args.proof_sigma}σ")
    print("-" * 60)
    print("MTDF VECTOR PILLARS:")
    for vr in mtdf_row.get("vector_results", []):
        exp_marker = " *" if vr.get("experimental") else ""
        print(f"  {vr['name']:25s} n={vr['n_data']:5d}  χ²={vr['chi2']:8.1f}  DOF={vr['dof']:5d}  χ²/ν={vr['chi2_red']:.4f}{exp_marker}")
    vec_chi2 = mtdf_row.get('vector_chi2', 0)
    vec_dof = mtdf_row.get('vector_dof', 1)
    print(f"  {'VECTOR TOTAL (prod)':25s}       χ²={vec_chi2:8.1f}  DOF={vec_dof:5d}  χ²/ν={vec_chi2/vec_dof:.4f}")
    print("  (* = experimental, excluded from totals)")
    print("-" * 60)
    print("MTDF TOTALS (STRICT vs DIAGNOSTIC):")
    # Strict totals exclude CMB* diagnostic pillar
    strict_chi2 = mtdf_row.get("combined_excl_cmb_chi2", float("nan"))
    strict_dof = mtdf_row.get("combined_excl_cmb_dof", 0)
    strict_red = mtdf_row.get("combined_excl_cmb_chi2_red", float("nan"))
    diag_chi2 = mtdf_row.get("cmb_chi2", 0.0)
    diag_dof = mtdf_row.get("cmb_dof", 0)
    diag_red = (diag_chi2 / diag_dof) if diag_dof else float("nan")
    all_chi2 = mtdf_row.get("combined_chi2", mtdf_row.get("chi2", float("nan")))
    all_dof = mtdf_row.get("combined_dof", mtdf_row.get("dof", 0))
    all_red = mtdf_row.get("combined_chi2_red", mtdf_row.get("chi2_red", float("nan")))

    print(f"  Strict combined χ² = {strict_chi2:.2f}  DOF = {strict_dof:5d}  χ²/ν = {strict_red:.4f}")
    if diag_dof:
        print(f"  CMB* diagnostic  χ² = {diag_chi2:.2f}  DOF = {diag_dof:5d}  χ²/ν = {diag_red:.4f}  (excluded from strict totals)")
    print(f"  Including diagnostics χ² = {all_chi2:.2f}  DOF = {all_dof:5d}  χ²/ν = {all_red:.4f}")
    if other_rows:
        print("-" * 60)
        print("Other models summary (scalar pillars only):")
        summ = []
        for r in other_rows:
            summ.append({
                "model": r["model"],
                "dof": r["dof"],
                "chi2_red": r["chi2_red"],
                "passes": r["passes"],
                "max_|z|": r["max_z"],
            })
        print(pd.DataFrame(summ).sort_values("chi2_red").to_string(index=False))
    # fσ₈ Growth Diagnostics - MTDF vs ΛCDM comparison
    print("-" * 60)
    print("fσ₈ GROWTH MODEL COMPARISON (MTDF vs ΛCDM):")
    fsig8_result = None
    for vr in mtdf_row.get("vector_results", []):
        if "fsig8" in vr.get("pillar_id", "").lower() or "growth" in vr.get("pillar_id", "").lower():
            fsig8_result = vr
            break

    if fsig8_result and "fsig8_diagnostics" in fsig8_result:
        diag = fsig8_result["fsig8_diagnostics"]
        obs = fsig8_result.get("fsig8_obs", diag["fsigma8"])
        pred_mtdf = diag["fsigma8"]

        # Get ΛCDM comparison
        z_f, fsig8_obs_f, cov_f = load_dr16_fsigma8(str(wb.parent))
        sigma_f = np.sqrt(np.diag(cov_f))

        # Fit ΛCDM with same data
        lcdm_params = {"H0": 70.0, "Omega_m": 0.3}
        pred_lcdm, lcdm_diag = lcdm_fsigma8_vector(
            z_f, lcdm_params, return_diagnostics=True,
            fit_sigma8=True, fsig8_obs=fsig8_obs_f, cov_matrix=cov_f
        )

        # MTDF results
        sigma8_mtdf = fsig8_result.get("sigma8_bf", diag.get("sigma8_bf"))
        sigma8_mtdf_err = fsig8_result.get("sigma8_err", diag.get("sigma8_err"))
        chi2_mtdf = fsig8_result['chi2']
        dof_mtdf = fsig8_result['dof']

        # ΛCDM results
        sigma8_lcdm = lcdm_diag.get("sigma8_bf")
        sigma8_lcdm_err = lcdm_diag.get("sigma8_err")
        chi2_lcdm = lcdm_diag.get("chi2_fitted")
        dof_lcdm = len(z_f) - 1

        print()
        print("  ┌─────────────────────────────────────────────────────────────┐")
        print("  │                   fσ₈ MODEL COMPARISON                      │")
        print("  ├───────────────────┬───────────────────┬─────────────────────┤")
        print("  │      Metric       │       MTDF        │        ΛCDM         │")
        print("  ├───────────────────┼───────────────────┼─────────────────────┤")
        print(f"  │ Best-fit σ₈,₀     │ {sigma8_mtdf:6.4f} ± {sigma8_mtdf_err:.4f} │ {sigma8_lcdm:6.4f} ± {sigma8_lcdm_err:.4f}   │")
        print(f"  │ χ²                │ {chi2_mtdf:17.2f} │ {chi2_lcdm:19.2f} │")
        print(f"  │ DOF               │ {dof_mtdf:17d} │ {dof_lcdm:19d} │")
        print(f"  │ χ²/ν              │ {chi2_mtdf/dof_mtdf:17.4f} │ {chi2_lcdm/dof_lcdm:19.4f} │")
        print("  └───────────────────┴───────────────────┴─────────────────────┘")
        print()
        print(f"  Planck 2018 reference: σ₈ = 0.811 ± 0.006")
        print(f"    MTDF tension with Planck: {(sigma8_mtdf - 0.811) / 0.006:+.1f}σ")
        print(f"    ΛCDM tension with Planck: {(sigma8_lcdm - 0.811) / 0.006:+.1f}σ")
        print()

        # Compute pulls for both models
        pulls_mtdf = (obs - pred_mtdf) / sigma_f
        pulls_lcdm = (obs - pred_lcdm) / sigma_f

        print("  Point-by-point comparison:")
        print("  " + "-" * 94)
        print(f"  {'z':>6s}  {'μ_MTDF':>8s}  {'f_MTDF':>8s}  {'fσ8_MTDF':>10s}  {'f_ΛCDM':>8s}  {'fσ8_ΛCDM':>10s}  {'fσ8_obs':>10s}  {'Δ_MTDF':>7s}  {'Δ_ΛCDM':>7s}")
        print("  " + "-" * 94)
        for i in range(len(diag["z"])):
            print(f"  {diag['z'][i]:6.3f}  {diag['mu_mtdf'][i]:8.4f}  {diag['f'][i]:8.4f}  {pred_mtdf[i]:10.4f}  "
                  f"{lcdm_diag['f'][i]:8.4f}  {pred_lcdm[i]:10.4f}  {obs[i]:10.4f}  {pulls_mtdf[i]:+7.2f}  {pulls_lcdm[i]:+7.2f}")
        print("  " + "-" * 94)
        print()
        print(f"  fσ₈ pillar summary (MTDF, included in combined totals):")
        print(f"    n = {fsig8_result['n_data']} DR16 points")
        print(f"    DOF = {fsig8_result['dof']} (n - 1 for fitted σ₈,₀)")
        print(f"    χ² = {fsig8_result['chi2']:.2f}")
        print(f"    χ²/ν = {fsig8_result['chi2_red']:.4f}")
        print()
        print(f"  μ(a) MTDF range over integration (a = 1e-3 to 1):")
        print(f"    min μ(a) = {diag['mu_min']:.6f}")
        print(f"    max μ(a) = {diag['mu_max']:.6f}")
    else:
        print("  [No fσ₈ diagnostics available]")

    print("=" * 60)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())

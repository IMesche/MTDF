# Author: Ingo Mesche
# Affiliation: Independent Researcher, Malta
# Framework: MTDF V74

"""
Central configuration for MTDF GPU Validation Pipeline.
Reads all parameters from DB_Workbook_STRICT_V18.xlsx.
"""

import numpy as np
from pathlib import Path

# ============================================================================
# PATHS
# ============================================================================

PROJECT_ROOT = Path(__file__).parent.parent  # repository root
SUBMISSION_DIR = PROJECT_ROOT / "validation"
DATA_DIR = SUBMISSION_DIR / "data"
WORKBOOK_PATH = DATA_DIR / "DB_Workbook_STRICT_V18.xlsx"
RESULTS_DIR = Path(__file__).parent / "results"

# Physical constants
C_LIGHT = 299792.458  # km/s
MPC_TO_M = 3.0857e22  # meters per Mpc


def load_workbook_params(workbook_path=None):
    """
    Read all MTDF parameters from the V18 workbook.
    Returns a flat dict of {token: value}.
    """
    import openpyxl

    if workbook_path is None:
        workbook_path = WORKBOOK_PATH

    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    params = {}

    # Read parameter sheets
    for sheet_name in ['Params_Fundamental', 'Params_Observational',
                       'Params_Coefficients', 'Params_Constants']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find column indices
        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        token_col = None
        value_col = None
        for i, h in enumerate(header):
            if 'input_token' in h or h == 'token' or h == 'parameter':
                token_col = i
            if 'value_si' in h or h == 'value' or 'value' in h:
                if value_col is None:  # take first value column
                    value_col = i

        if token_col is None or value_col is None:
            continue

        for row in rows[1:]:
            if row[token_col] is None:
                continue
            token = str(row[token_col]).strip()
            val = row[value_col]
            if val is not None:
                try:
                    params[token] = float(val)
                except (ValueError, TypeError):
                    params[token] = val

    wb.close()
    return params


def build_mtdf_params(workbook_params=None):
    """
    Build the MTDF parameter dict used by all physics functions.
    Mirrors the parameter hierarchy in run_validate.py.
    """
    if workbook_params is None:
        workbook_params = load_workbook_params()

    def get(key, *alternates, default=None):
        """Get parameter by key, trying alternates."""
        for k in [key] + list(alternates):
            # Try exact match
            if k in workbook_params:
                return workbook_params[k]
            # Try case-insensitive
            for wk, wv in workbook_params.items():
                if wk.lower().replace(' ', '').replace('_', '') == k.lower().replace(' ', '').replace('_', ''):
                    return wv
        return default

    H0 = get('H0', 'h0', 'hubble_constant', default=70.0)
    h = H0 / 100.0

    omegab_h2 = get('omegab_h2', 'omega_b_h2', default=0.02236)
    omegam_h2 = get('omegam_h2', 'omega_m_h2', default=0.1430)

    Omega_b = omegab_h2 / h**2
    Omega_m = omegam_h2 / h**2

    params = {
        'H0': H0,
        'h': h,
        'omegab_h2': omegab_h2,
        'omegam_h2': omegam_h2,
        'Omega_b': Omega_b,
        'Omega_m': Omega_m,
        'kappa': get('kappa', 'kappa_spec', default=1.02e-3),  # Anchor: approx f_kick/3
        'alpha': get('alpha', default=1.30),
        'beta_eos': get('beta_eos', 'betaeos', default=0.573),
        'z_t': get('z_t', 'zt', default=0.74),
        'sigma8': get('sigma8', 'sigma_8', default=0.811),
    }

    # Derived MTDF quantities
    params['RS_ZSTAR_RATIO'] = get('RS_ZSTAR_RATIO', default=0.9819)

    return params
